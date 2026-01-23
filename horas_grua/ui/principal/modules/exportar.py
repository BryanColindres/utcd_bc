import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from copy import copy
from datetime import datetime
import locale
import sys
import os
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# IMPORTS DE TU CAPA DE DATOS
# ------------------------------------------------------------
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../../..")))
from db_controller import obtener_proveedor_codigo, obtener_datos_sectores

# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

SHEET_NAME = "V2"
START_ROW = 14
MAX_TEMPLATE_ROWS = 15
def resource_path(relative_path):
    """
    Obtiene la ruta correcta tanto en desarrollo como en .exe
    """
    try:
        base_path = sys._MEIPASS  # PyInstaller
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

TEMPLATE_FILENAME = resource_path(
    os.path.join("resources", "plantilla_control_servicios.xlsx")
)


# ============================================================
# LOCALE PARA FECHA EN ESPAÑOL
# ============================================================

try:
    locale.setlocale(locale.LC_TIME, "spanish")
except:
    locale.setlocale(locale.LC_TIME, "es_ES")

# ============================================================
# UTILIDADES
# ============================================================

def fecha_literal_es(fecha):
    fecha = pd.to_datetime(fecha)
    return fecha.strftime("%#d de %B de %Y")

def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        if src.has_style:
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.fill = copy(src.fill)
            tgt.number_format = copy(src.number_format)
            tgt.alignment = copy(src.alignment)

def write_merged(ws, cell_ref, value):
    cell = ws[cell_ref]

    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            anchor = ws.cell(merged.min_row, merged.min_col)
            anchor.value = value
            return

    raise RuntimeError(f"No se encontró rango combinado para {cell_ref}")

# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================

def export_to_excel(data, columns=None, id_sector=None):
    print("Iniciando exportación a Excel...")
    print(columns)
    try:
        # ----------------------------------------------------
        # VALIDACIÓN
        # ----------------------------------------------------
        if data is None or len(data) == 0:
            messagebox.showerror("Error", "No se encontraron datos para exportar.")
            return

        # ----------------------------------------------------
        # DATAFRAME
        # ----------------------------------------------------
        if isinstance(data, pd.DataFrame):
            df = data.copy()
        else:
            n_cols = len(data[0])
            if columns is None:
                columns = [f"Columna_{i+1}" for i in range(n_cols)]
            else:
                columns = columns[:n_cols]
            df = pd.DataFrame(data, columns=columns)


        total_rows = len(df)

        # ----------------------------------------------------
        # DATOS DE ENCABEZADO (PRIMERO)
        # ----------------------------------------------------
        fecha = fecha_literal_es(datetime.now())
        supervisor, centro_costo, sector = obtener_datos_sectores(id_sector)
        print(df.columns)
        orden_compra = df["ORDEN_COMPRA"].dropna().unique()[0]
        equipo = df["TIPO_EQUIPO"].dropna().unique()[0]
        proveedor = obtener_proveedor_codigo(orden_compra,equipo)
        df = df.drop(columns=["ID", "ORDEN_COMPRA", "TIPO_EQUIPO"], errors="ignore")

        # ----------------------------------------------------
        # ARCHIVO DESTINO
        # ----------------------------------------------------
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if not file_path:
            return

        # ----------------------------------------------------
        # CARGAR PLANTILLA
        # ----------------------------------------------------
        wb = load_workbook(TEMPLATE_FILENAME)
        ws = wb[SHEET_NAME]

        # ----------------------------------------------------
        # ESCRIBIR ENCABEZADO (C:J)
        # ----------------------------------------------------
        write_merged(ws, "C4", fecha)
        write_merged(ws, "C5", sector)
        write_merged(ws, "C6", supervisor)
        write_merged(ws, "C7", orden_compra)
        write_merged(ws, "C8", centro_costo)
        write_merged(ws, "C9", proveedor)

        # ----------------------------------------------------
        # FILAS DINÁMICAS
        # ----------------------------------------------------
        if total_rows > MAX_TEMPLATE_ROWS:
            extra = total_rows - MAX_TEMPLATE_ROWS
            footer_start = START_ROW + MAX_TEMPLATE_ROWS
            ws.insert_rows(footer_start, extra)

            for i in range(extra):
                copy_row_style(ws, footer_start - 1, footer_start + i)

        # ----------------------------------------------------
        # ESCRITURA DE REGISTROS (SEGURA)
        # ----------------------------------------------------
        for r_idx, row in enumerate(df.itertuples(index=False), start=START_ROW):
            for c_idx, val in enumerate(row, start=2):  # ← empieza en columna B
                col_letter = get_column_letter(c_idx)
                write_merged(ws, f"{col_letter}{r_idx}", val)


        # ----------------------------------------------------
        # GUARDAR
        # ----------------------------------------------------
        wb.save(file_path)

        messagebox.showinfo(
            "Éxito",
            f"Archivo exportado correctamente:\n{file_path}"
        )

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error inesperado:\n{e}")

